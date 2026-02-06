import argparse
import hashlib
import json
import os
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from llama_index.core import Document, StorageContext, VectorStoreIndex

from doc_index_common import (
    DEFAULT_OCR_LANGS,
    compute_sha256,
    configure_embedding,
    detect_platform,
    install_ocr_tools,
    is_binary_file,
    is_tool_available,
    load_docs_cache,
    run_cmd,
    save_docs_cache,
)

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}
PDF_EXTS = {".pdf"}
XLSX_EXTS = {".xlsx"}
PDF_OCR_ENGINES = {"auto", "legacy", "ocrmypdf"}
PDF_CLEAN_MODES = {"none", "safe", "aggressive"}
IGNORED_DIRS = {
    ".git",
    ".codex",
    "node_modules",
    ".venv",
    "venv",
    "__pycache__",
    "dist",
    "build",
}
IGNORED_EXTS = {
    ".exe",
    ".dll",
    ".so",
    ".dylib",
    ".bin",
    ".dat",
    ".class",
    ".jar",
    ".pyc",
    ".o",
    ".a",
    ".zip",
    ".tar",
    ".gz",
    ".7z",
    ".rar",
    ".sqlite",
    ".db",
}
MAX_LINES_PER_DOC = 200
PDF_OCR_DPI = 300
PDF_LOW_TEXT_THRESHOLD = 20


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build or update the Doc Master index.")
    parser.add_argument("--root", default=os.getcwd(), help="Workspace root (default: cwd).")
    parser.add_argument(
        "--index-dir",
        default=None,
        help="Index directory (default: <root>/.codex/doc-master).",
    )
    parser.add_argument(
        "--languages",
        default=None,
        help="OCR languages (default: ita+eng).",
    )
    parser.add_argument(
        "--include-hidden",
        action="store_true",
        help="Include hidden files and folders.",
    )
    parser.add_argument(
        "--force-rebuild",
        action="store_true",
        help="Rebuild index even if no changes are detected.",
    )
    parser.add_argument(
        "--embed-model",
        default=None,
        help="HuggingFace embedding model id.",
    )
    parser.add_argument(
        "--pdf-ocr-engine",
        choices=sorted(PDF_OCR_ENGINES),
        default=None,
        help="PDF OCR engine: auto, legacy, or ocrmypdf.",
    )
    parser.add_argument(
        "--pdf-clean",
        choices=sorted(PDF_CLEAN_MODES),
        default=None,
        help="PDF cleanup mode when using OCRmyPDF.",
    )
    parser.add_argument(
        "--pdf-unpaper-args",
        default=None,
        help="Arguments passed to OCRmyPDF via --unpaper-args.",
    )
    parser.add_argument(
        "--no-auto-install-ocr-tools",
        action="store_true",
        help="Disable automatic installation of missing OCRmyPDF/unpaper tools.",
    )
    return parser.parse_args()


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def env_flag(name: str, default: bool) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    value = raw.strip().lower()
    if value in {"1", "true", "yes", "on"}:
        return True
    if value in {"0", "false", "no", "off"}:
        return False
    return default


def validate_choice(value: str, valid: set, label: str) -> str:
    normalized = value.strip().lower()
    if normalized not in valid:
        allowed = ", ".join(sorted(valid))
        raise SystemExit(f"Invalid {label}: {value}. Allowed values: {allowed}.")
    return normalized


def is_under(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def storage_ready(storage_dir: Path) -> bool:
    required = (
        "docstore.json",
        "index_store.json",
        "default__vector_store.json",
    )
    return all((storage_dir / name).exists() for name in required)


def should_skip_dir(path: Path, index_dir: Path, include_hidden: bool) -> bool:
    name = path.name
    if name in IGNORED_DIRS:
        return True
    if not include_hidden and name.startswith("."):
        return True
    if is_under(path, index_dir):
        return True
    return False


def should_skip_file(path: Path, include_hidden: bool) -> bool:
    name = path.name
    if not include_hidden and name.startswith("."):
        return True
    if path.suffix.lower() in IGNORED_EXTS:
        return True
    return False


def collect_files(root: Path, index_dir: Path, include_hidden: bool) -> List[Path]:
    files: List[Path] = []
    for dirpath, dirnames, filenames in os.walk(root):
        dir_path = Path(dirpath)
        dirnames[:] = [
            name
            for name in dirnames
            if not should_skip_dir(dir_path / name, index_dir, include_hidden)
        ]
        for name in filenames:
            path = dir_path / name
            if should_skip_file(path, include_hidden):
                continue
            files.append(path)
    files.sort()
    return files


def read_text_file(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="utf-8", errors="ignore")


def extract_text_docs(text: str, rel_path: str, file_type: str) -> List[Document]:
    docs: List[Document] = []
    lines = text.splitlines()
    if not lines:
        return docs
    for start in range(0, len(lines), MAX_LINES_PER_DOC):
        chunk_lines = lines[start : start + MAX_LINES_PER_DOC]
        chunk_text = "\n".join(chunk_lines).strip()
        if not chunk_text:
            continue
        line_range = f"{start + 1}-{start + len(chunk_lines)}"
        metadata = {
            "file_path": rel_path,
            "file_type": file_type,
            "line_range": line_range,
        }
        docs.append(Document(text=chunk_text, metadata=metadata))
    return docs


def ensure_ocr_ready() -> Tuple[object, object]:
    try:
        import pytesseract
    except Exception as exc:
        raise RuntimeError("Missing pytesseract. Install it to enable OCR.") from exc
    try:
        from PIL import Image
    except Exception as exc:
        raise RuntimeError("Missing Pillow. Install it to enable OCR.") from exc
    return pytesseract, Image


def ocr_image(image, languages: str) -> str:
    pytesseract, _ = ensure_ocr_ready()
    return pytesseract.image_to_string(image, lang=languages)


def extract_pdf_docs_legacy(
    path: Path,
    rel_path: str,
    languages: str,
    ocr_method: str,
) -> List[Document]:
    try:
        import fitz
    except Exception as exc:
        raise RuntimeError("Missing PyMuPDF (pymupdf). Install it to read PDFs.") from exc

    docs: List[Document] = []
    with fitz.open(path) as pdf_doc:
        for index in range(pdf_doc.page_count):
            page = pdf_doc.load_page(index)
            text = page.get_text("text") or ""
            method = "text"
            if len(text.strip()) < PDF_LOW_TEXT_THRESHOLD:
                _, Image = ensure_ocr_ready()
                pix = page.get_pixmap(dpi=PDF_OCR_DPI)
                mode = "RGB" if pix.alpha == 0 else "RGBA"
                image = Image.frombytes(mode, (pix.width, pix.height), pix.samples)
                text = ocr_image(image, languages)
                method = ocr_method
            if not text.strip():
                continue
            metadata = {
                "file_path": rel_path,
                "file_type": "pdf",
                "page_number": index + 1,
                "extract_method": method,
            }
            docs.append(Document(text=text, metadata=metadata))
    return docs


def extract_pdf_text_layer_docs(path: Path, rel_path: str, method: str) -> List[Document]:
    try:
        import fitz
    except Exception as exc:
        raise RuntimeError("Missing PyMuPDF (pymupdf). Install it to read PDFs.") from exc

    docs: List[Document] = []
    with fitz.open(path) as pdf_doc:
        for index in range(pdf_doc.page_count):
            text = (pdf_doc.load_page(index).get_text("text") or "").strip()
            if not text:
                continue
            metadata = {
                "file_path": rel_path,
                "file_type": "pdf",
                "page_number": index + 1,
                "extract_method": method,
            }
            docs.append(Document(text=text, metadata=metadata))
    return docs


def pdf_has_low_text_pages(path: Path) -> bool:
    try:
        import fitz
    except Exception as exc:
        raise RuntimeError("Missing PyMuPDF (pymupdf). Install it to read PDFs.") from exc

    with fitz.open(path) as pdf_doc:
        for index in range(pdf_doc.page_count):
            text = pdf_doc.load_page(index).get_text("text") or ""
            if len(text.strip()) < PDF_LOW_TEXT_THRESHOLD:
                return True
    return False


def build_ocrmypdf_command(
    input_pdf: Path,
    output_pdf: Path,
    languages: str,
    clean_mode: str,
    unpaper_args: str,
    use_unpaper: bool,
) -> List[str]:
    command: List[str] = [
        "ocrmypdf",
        "--skip-text",
        "--rotate-pages",
        "--deskew",
        "-l",
        languages,
    ]
    if clean_mode in {"safe", "aggressive"} and use_unpaper:
        command.append("--clean")
        if clean_mode == "aggressive":
            command.append("--clean-final")
        if unpaper_args.strip():
            command.extend(["--unpaper-args", unpaper_args.strip()])
    command.extend([str(input_pdf), str(output_pdf)])
    return command


def extract_pdf_docs_ocrmypdf(
    path: Path,
    rel_path: str,
    languages: str,
    clean_mode: str,
    unpaper_args: str,
    use_unpaper: bool,
    scratch_dir: Path,
) -> Tuple[Optional[List[Document]], Optional[str]]:
    scratch_dir.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix="ocrmypdf-", suffix=".pdf", dir=str(scratch_dir))
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        command = build_ocrmypdf_command(
            path,
            temp_path,
            languages,
            clean_mode,
            unpaper_args,
            use_unpaper,
        )
        command_result = run_cmd(command, timeout=7200, check=False)
        if not command_result.get("ok"):
            reason = (
                str(command_result.get("stderr") or "").strip()
                or str(command_result.get("stdout") or "").strip()
                or "ocrmypdf failed"
            )
            return None, reason

        docs = extract_pdf_text_layer_docs(temp_path, rel_path, "ocrmypdf")
        if not docs:
            return None, "ocrmypdf produced no text"
        return docs, None
    finally:
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass


def extract_pdf_docs(
    path: Path,
    rel_path: str,
    languages: str,
    pdf_engine: str,
    pdf_clean: str,
    pdf_unpaper_args: str,
    ocr_tools: Dict[str, Any],
    scratch_dir: Path,
    log_entries: List[Dict[str, Any]],
) -> List[Document]:
    if pdf_engine == "legacy":
        return extract_pdf_docs_legacy(path, rel_path, languages, "ocr-legacy")

    try:
        has_low_text = pdf_has_low_text_pages(path)
    except Exception as exc:
        log_entries.append(
            {
                "path": rel_path,
                "status": "warning",
                "reason": f"PDF scan detection failed: {exc}",
            }
        )
        return extract_pdf_docs_legacy(path, rel_path, languages, "ocr-legacy")

    if pdf_engine == "auto" and not has_low_text:
        return extract_pdf_docs_legacy(path, rel_path, languages, "ocr-legacy")

    if not ocr_tools.get("ocrmypdf_available"):
        log_entries.append(
            {
                "path": rel_path,
                "status": "warning",
                "reason": "ocrmypdf unavailable; using legacy OCR fallback",
            }
        )
        return extract_pdf_docs_legacy(path, rel_path, languages, "ocrmypdf-fallback-legacy")

    use_unpaper = pdf_clean != "none" and bool(ocr_tools.get("unpaper_available"))
    if pdf_clean != "none" and not use_unpaper:
        log_entries.append(
            {
                "path": rel_path,
                "status": "warning",
                "reason": "unpaper unavailable; OCRmyPDF clean step disabled",
            }
        )

    docs, failure_reason = extract_pdf_docs_ocrmypdf(
        path,
        rel_path,
        languages,
        pdf_clean,
        pdf_unpaper_args,
        use_unpaper,
        scratch_dir,
    )
    if docs:
        return docs

    if failure_reason:
        log_entries.append(
            {
                "path": rel_path,
                "status": "warning",
                "reason": f"ocrmypdf failed ({failure_reason}); using legacy OCR fallback",
            }
        )
    return extract_pdf_docs_legacy(path, rel_path, languages, "ocrmypdf-fallback-legacy")


def extract_image_docs(path: Path, rel_path: str, languages: str) -> List[Document]:
    _, Image = ensure_ocr_ready()
    with Image.open(path) as image:
        text = ocr_image(image, languages)
    if not text.strip():
        return []
    metadata = {"file_path": rel_path, "file_type": "image"}
    return [Document(text=text, metadata=metadata)]


def extract_xlsx_docs(path: Path, rel_path: str) -> List[Document]:
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("Missing openpyxl. Install it to read XLSX files.") from exc
    from openpyxl.utils.cell import get_column_letter, range_boundaries

    docs: List[Document] = []
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        dimension = sheet.calculate_dimension()
        if dimension == "A1" and sheet["A1"].value is None:
            continue
        min_col, min_row, max_col, max_row = range_boundaries(dimension)
        col_start = get_column_letter(min_col)
        col_end = get_column_letter(max_col)
        row = min_row
        while row <= max_row:
            row_end = min(row + 49, max_row)
            lines = []
            for values in sheet.iter_rows(
                min_row=row,
                max_row=row_end,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            ):
                line = "\t".join("" if value is None else str(value) for value in values)
                if line.strip():
                    lines.append(line)
            text = "\n".join(lines).strip()
            if text:
                cell_range = f"{col_start}{row}:{col_end}{row_end}"
                metadata = {
                    "file_path": rel_path,
                    "file_type": "xlsx",
                    "sheet": sheet.title,
                    "cell_range": cell_range,
                }
                docs.append(Document(text=text, metadata=metadata))
            row = row_end + 1
    workbook.close()
    return docs


def extract_docs_for_file(
    path: Path,
    rel_path: str,
    languages: str,
    pdf_options: Dict[str, Any],
    ocr_tools: Dict[str, Any],
    log_entries: List[Dict[str, Any]],
) -> Tuple[List[Document], str]:
    ext = path.suffix.lower()
    if ext in PDF_EXTS:
        return (
            extract_pdf_docs(
                path,
                rel_path,
                languages,
                str(pdf_options["engine"]),
                str(pdf_options["clean"]),
                str(pdf_options["unpaper_args"]),
                ocr_tools,
                Path(pdf_options["scratch_dir"]),
                log_entries,
            ),
            "pdf",
        )
    if ext in IMAGE_EXTS:
        return extract_image_docs(path, rel_path, languages), "image"
    if ext in XLSX_EXTS:
        return extract_xlsx_docs(path, rel_path), "xlsx"
    if is_binary_file(path):
        return [], "binary"
    text = read_text_file(path)
    return extract_text_docs(text, rel_path, "text"), "text"


def build_ocr_config_fingerprint(
    languages: str,
    pdf_engine: str,
    pdf_clean: str,
    pdf_unpaper_args: str,
) -> str:
    payload = {
        "languages": languages,
        "pdf_engine": pdf_engine,
        "pdf_clean": pdf_clean,
        "pdf_unpaper_args": pdf_unpaper_args,
    }
    encoded = json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":")).encode(
        "utf-8"
    )
    return hashlib.sha256(encoded).hexdigest()[:16]


def summarize_install_result(raw_result: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not raw_result:
        return None

    tool_details: Dict[str, Dict[str, Any]] = {}
    for tool_name, state in (raw_result.get("tools") or {}).items():
        tool_details[tool_name] = {
            "status": state.get("status"),
            "reason": state.get("reason"),
            "available_before": bool(state.get("available_before")),
            "available_after": bool(state.get("available_after")),
        }

    return {
        "platform": raw_result.get("platform"),
        "attempted": bool(raw_result.get("attempted")),
        "package_manager": raw_result.get("package_manager"),
        "success": bool(raw_result.get("success")),
        "full_pipeline_available": bool(raw_result.get("full_pipeline_available")),
        "tools": tool_details,
    }


def load_manifest(path: Path) -> Dict[str, object]:
    if not path.exists():
        return {"version": 1, "files": {}}
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def save_manifest(path: Path, manifest: Dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(manifest, handle, ensure_ascii=True, indent=2)


def write_index_summary(
    path: Path,
    root: Path,
    index_dir: Path,
    languages: str,
    stats: Dict[str, int],
    doc_count: int,
    last_log: str,
    embed_info: Dict[str, str],
    ocr_config: Dict[str, Any],
    ocr_tools: Dict[str, Any],
) -> None:
    lines = [
        "# Doc Master",
        "",
        f"Root: {root}",
        f"Index dir: {index_dir}",
        f"Last run: {now_iso()}",
        f"OCR languages: {languages}",
        f"PDF OCR engine: {ocr_config.get('pdf_engine')}",
        f"PDF clean mode: {ocr_config.get('pdf_clean')}",
        f"Auto-install OCR tools: {ocr_config.get('auto_install_ocr_tools')}",
        f"OCRmyPDF available: {ocr_tools.get('ocrmypdf', {}).get('available')}",
        f"unpaper available: {ocr_tools.get('unpaper', {}).get('available')}",
        f"Embedding: {embed_info.get('type')}:{embed_info.get('model')}",
        "",
        "Counts:",
        f"- total files: {stats.get('total_files', 0)}",
        f"- processed: {stats.get('processed', 0)}",
        f"- skipped: {stats.get('skipped', 0)}",
        f"- errors: {stats.get('errors', 0)}",
        f"- removed: {stats.get('removed', 0)}",
        f"- documents: {doc_count}",
        "",
        "Locations:",
        f"- storage: {index_dir / 'storage'}",
        f"- cache: {index_dir / 'cache'}",
        f"- logs: {index_dir / 'logs'}",
        "",
        f"Last log: {last_log}",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    args = parse_args()
    root = Path(args.root).resolve()
    index_dir = Path(args.index_dir).resolve() if args.index_dir else root / ".codex" / "doc-master"
    index_dir = index_dir.resolve()

    if not is_under(index_dir, root):
        raise SystemExit("Index directory must be inside the workspace root.")

    languages = args.languages or os.getenv("DOC_INDEX_OCR_LANGS", DEFAULT_OCR_LANGS)
    pdf_engine_raw = args.pdf_ocr_engine or os.getenv("DOC_INDEX_PDF_OCR_ENGINE", "auto")
    pdf_clean_raw = args.pdf_clean or os.getenv("DOC_INDEX_PDF_CLEAN", "safe")
    pdf_unpaper_args = args.pdf_unpaper_args
    if pdf_unpaper_args is None:
        pdf_unpaper_args = os.getenv("DOC_INDEX_PDF_UNPAPER_ARGS", "")
    pdf_ocr_engine = validate_choice(pdf_engine_raw, PDF_OCR_ENGINES, "PDF OCR engine")
    pdf_clean = validate_choice(pdf_clean_raw, PDF_CLEAN_MODES, "PDF clean mode")

    auto_install_ocr_tools = (
        not args.no_auto_install_ocr_tools
        and env_flag("DOC_INDEX_AUTO_INSTALL_OCR_TOOLS", True)
    )

    index_dir.mkdir(parents=True, exist_ok=True)
    cache_dir = index_dir / "cache"
    log_dir = index_dir / "logs"
    storage_dir = index_dir / "storage"
    scratch_ocr_dir = index_dir / "scratch" / "tmp" / "ocrmypdf"
    cache_dir.mkdir(parents=True, exist_ok=True)
    log_dir.mkdir(parents=True, exist_ok=True)
    storage_dir.mkdir(parents=True, exist_ok=True)
    scratch_ocr_dir.mkdir(parents=True, exist_ok=True)

    embed_info = configure_embedding(args.embed_model)

    ocr_fingerprint = build_ocr_config_fingerprint(
        languages,
        pdf_ocr_engine,
        pdf_clean,
        pdf_unpaper_args,
    )

    manifest_path = index_dir / "manifest.json"
    manifest = load_manifest(manifest_path)
    previous_files: Dict[str, Dict[str, object]] = manifest.get("files", {})  # type: ignore[assignment]

    log_entries: List[Dict[str, Any]] = []
    stats = {"total_files": 0, "processed": 0, "skipped": 0, "errors": 0, "removed": 0}
    documents: List[Document] = []
    new_files: Dict[str, Dict[str, object]] = {}
    reprocessed = 0

    files = collect_files(root, index_dir, args.include_hidden)
    pdf_files = [path for path in files if path.suffix.lower() in PDF_EXTS]
    has_pdf_files = bool(pdf_files)

    platform_name = detect_platform()
    ocr_tool_state: Dict[str, Any] = {
        "platform": platform_name,
        "ocrmypdf_available": is_tool_available("ocrmypdf"),
        "unpaper_available": is_tool_available("unpaper"),
    }
    install_attempt_at: Optional[str] = None
    install_summary: Optional[Dict[str, Any]] = None

    pdfs_need_ocr = False
    if has_pdf_files:
        if pdf_ocr_engine == "ocrmypdf":
            pdfs_need_ocr = True
        elif pdf_ocr_engine == "auto":
            for pdf_path in pdf_files:
                try:
                    if pdf_has_low_text_pages(pdf_path):
                        pdfs_need_ocr = True
                        break
                except Exception as exc:
                    log_entries.append(
                        {
                            "path": str(pdf_path.relative_to(root).as_posix()),
                            "status": "warning",
                            "reason": f"Failed to pre-scan PDF text layer: {exc}",
                        }
                    )
                    # If detection fails, keep conservative behavior and allow OCR tool setup.
                    pdfs_need_ocr = True
                    break

    if has_pdf_files and pdf_ocr_engine != "legacy":
        needs_ocrmypdf = pdfs_need_ocr
        needs_unpaper = needs_ocrmypdf and pdf_clean != "none" and platform_name != "windows"
        missing_required_tool = (
            (needs_ocrmypdf and not ocr_tool_state["ocrmypdf_available"])
            or (needs_unpaper and not ocr_tool_state["unpaper_available"])
        )

        if missing_required_tool:
            if auto_install_ocr_tools:
                install_attempt_at = now_iso()
                raw_install = install_ocr_tools(platform_name=platform_name, allow_sudo=True)
                install_summary = summarize_install_result(raw_install)
                ocr_tool_state["ocrmypdf_available"] = is_tool_available("ocrmypdf")
                ocr_tool_state["unpaper_available"] = is_tool_available("unpaper")
                install_ok = bool(raw_install.get("success"))
                log_entries.append(
                    {
                        "status": "info" if install_ok else "warning",
                        "event": "ocr_tool_install",
                        "reason": "installed" if install_ok else "failed",
                        "details": install_summary,
                    }
                )
            else:
                log_entries.append(
                    {
                        "status": "warning",
                        "event": "ocr_tool_install",
                        "reason": "missing tools and auto-install disabled",
                    }
                )

    pdf_options: Dict[str, Any] = {
        "engine": pdf_ocr_engine,
        "clean": pdf_clean,
        "unpaper_args": pdf_unpaper_args,
        "scratch_dir": scratch_ocr_dir,
    }

    for path in files:
        rel_path = path.relative_to(root).as_posix()
        stats["total_files"] += 1
        prev = previous_files.get(rel_path)
        stat = path.stat()
        prev_cache_path = None
        if prev:
            prev_cache_rel = prev.get("cache_path")
            if prev_cache_rel:
                candidate = index_dir / str(prev_cache_rel)
                if candidate.is_file():
                    prev_cache_path = candidate
        if (
            prev
            and prev.get("mtime") == stat.st_mtime
            and prev.get("size") == stat.st_size
            and prev.get("ocr_fingerprint") == ocr_fingerprint
            and prev_cache_path
        ):
            cache_docs = load_docs_cache(prev_cache_path)
            documents.extend(cache_docs)
            log_entries.append(
                {"path": rel_path, "status": "skipped", "reason": "unchanged"}
            )
            stats["skipped"] += 1
            updated = dict(prev)
            updated["doc_count"] = len(cache_docs)
            new_files[rel_path] = updated
            continue

        file_hash = compute_sha256(path)
        new_cache_path = cache_dir / f"{file_hash}-{ocr_fingerprint}.json"
        if (
            prev
            and prev.get("hash") == file_hash
            and prev.get("ocr_fingerprint") == ocr_fingerprint
            and new_cache_path.exists()
        ):
            cache_docs = load_docs_cache(new_cache_path)
            documents.extend(cache_docs)
            log_entries.append(
                {"path": rel_path, "status": "skipped", "reason": "hash match"}
            )
            stats["skipped"] += 1
            new_files[rel_path] = {
                **prev,
                "mtime": stat.st_mtime,
                "size": stat.st_size,
                "cache_path": str(new_cache_path.relative_to(index_dir)),
                "doc_count": len(cache_docs),
                "ocr_fingerprint": ocr_fingerprint,
            }
            continue

        try:
            docs, file_type = extract_docs_for_file(
                path,
                rel_path,
                languages,
                pdf_options,
                ocr_tool_state,
                log_entries,
            )
            save_docs_cache(new_cache_path, docs, rel_path, file_type)
            reprocessed += 1
            if docs:
                documents.extend(docs)
                stats["processed"] += 1
                reason = "extracted"
            else:
                stats["skipped"] += 1
                reason = "binary" if file_type == "binary" else "no text"
            log_entries.append(
                {
                    "path": rel_path,
                    "status": "processed" if docs else "skipped",
                    "reason": reason,
                    "file_type": file_type,
                }
            )
            new_files[rel_path] = {
                "hash": file_hash,
                "mtime": stat.st_mtime,
                "size": stat.st_size,
                "cache_path": str(new_cache_path.relative_to(index_dir)),
                "file_type": file_type,
                "doc_count": len(docs),
                "ocr_fingerprint": ocr_fingerprint,
            }
        except Exception as exc:
            if prev and prev_cache_path:
                cache_docs = load_docs_cache(prev_cache_path)
                documents.extend(cache_docs)
                updated = dict(prev)
                updated["doc_count"] = len(cache_docs)
                new_files[rel_path] = updated
                log_entries.append(
                    {
                        "path": rel_path,
                        "status": "error",
                        "reason": str(exc),
                        "used_cache": True,
                    }
                )
            else:
                log_entries.append(
                    {"path": rel_path, "status": "error", "reason": str(exc)}
                )
            stats["errors"] += 1

    removed = set(previous_files.keys()) - set(new_files.keys())
    for rel_path in sorted(removed):
        stats["removed"] += 1
        log_entries.append(
            {"path": rel_path, "status": "skipped", "reason": "removed"}
        )

    changed = reprocessed > 0 or stats["removed"] > 0
    storage_ok = storage_ready(storage_dir)
    if documents and (changed or args.force_rebuild or not storage_ok):
        storage_context = StorageContext.from_defaults()
        VectorStoreIndex.from_documents(documents, storage_context=storage_context)
        storage_context.persist(persist_dir=str(storage_dir))
    elif not documents:
        log_entries.append(
            {"status": "error", "reason": "no documents extracted; index not rebuilt"}
        )

    log_name = f"index-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.jsonl"
    log_path = log_dir / log_name
    with log_path.open("w", encoding="utf-8") as handle:
        for entry in log_entries:
            handle.write(json.dumps(entry, ensure_ascii=True) + "\n")

    manifest["root"] = str(root)
    manifest["index_dir"] = str(index_dir)
    manifest["ocr_languages"] = languages
    manifest["last_run"] = now_iso()
    manifest["embed"] = embed_info
    manifest["files"] = new_files
    manifest["stats"] = stats
    manifest["ocr"] = {
        "languages": languages,
        "pdf_engine": pdf_ocr_engine,
        "pdf_clean": pdf_clean,
        "pdf_unpaper_args": pdf_unpaper_args,
        "config_fingerprint": ocr_fingerprint,
        "auto_install_ocr_tools": auto_install_ocr_tools,
    }
    manifest["ocr_tools"] = {
        "platform": platform_name,
        "ocrmypdf": {"available": bool(ocr_tool_state["ocrmypdf_available"])},
        "unpaper": {"available": bool(ocr_tool_state["unpaper_available"])},
        "last_install_attempt": install_attempt_at,
        "last_install_result": install_summary,
    }
    save_manifest(manifest_path, manifest)

    write_index_summary(
        index_dir / "INDEX.md",
        root,
        index_dir,
        languages,
        stats,
        len(documents),
        str(log_path),
        embed_info,
        manifest["ocr"],
        manifest["ocr_tools"],
    )

    return 0


if __name__ == "__main__":
    sys.exit(main())
